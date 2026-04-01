from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


EXPECTED_RECEIPT_FILES = 4
OUTPUT_NAME = "validacao_consolidada_financeiro.xlsx"


@dataclass
class SourcePaths:
    desktop: Path
    cashbook: Path
    expense_report: Path
    receipt_reports: list[Path]
    output_workspace: Path
    output_desktop: Path


def clean_text(value: object) -> str | None:
    if value is None:
        return None
    text = " ".join(str(value).replace("\xa0", " ").split())
    return text.replace("ş", "º")


def norm_key(value: object) -> str:
    text = clean_text(value) or ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def parse_date_br(value: object) -> pd.Timestamp | pd.NaT:
    text = clean_text(value)
    if not text or text == "-":
        return pd.NaT
    try:
        return pd.Timestamp(datetime.strptime(text, "%d/%m/%y").date())
    except ValueError:
        try:
            return pd.to_datetime(text)
        except Exception:
            return pd.NaT


def parse_currency(value: object) -> float | None:
    text = clean_text(value)
    if not text or text == "-":
        return None
    text = text.replace("R$", "").replace(".", "").replace(",", ".").strip()
    try:
        return float(text)
    except ValueError:
        return None


def split_pipe(value: object) -> tuple[str | None, str | None]:
    text = clean_text(value)
    if not text or text == "-":
        return None, None
    left, _, right = text.partition("|")
    left = left.strip() or None
    right = right.strip() or None
    return left, right


def detect_paths() -> SourcePaths:
    base = Path.home() / "OneDrive"
    desktop = next(path for path in base.iterdir() if "Trabalho" in path.name)
    cashbook = desktop / "livro caixa.xlsx"
    expense_report = desktop / "FaturaspagasporPeriodo.xls"
    receipt_reports = sorted(desktop.glob("FaturasrecebidasporPeriodo*.xls"))
    workspace = Path.cwd()
    return SourcePaths(
        desktop=desktop,
        cashbook=cashbook,
        expense_report=expense_report,
        receipt_reports=receipt_reports,
        output_workspace=workspace / OUTPUT_NAME,
        output_desktop=desktop / OUTPUT_NAME,
    )


def parse_html_report(path: Path, kind: str) -> pd.DataFrame:
    html = path.read_bytes().decode("utf-8", errors="ignore")
    soup = BeautifulSoup(html, "html.parser")

    current_category: str | None = None
    header_keys: list[str] | None = None
    rows: list[dict[str, object]] = []

    for tr in soup.find_all("tr"):
        texts = [clean_text(td.get_text(" ", strip=True)) for td in tr.find_all(["th", "td"])]
        texts = [text for text in texts if text is not None]
        if not any(texts):
            continue

        if texts[0] == "Vencto.":
            header_keys = [norm_key(value) for value in texts]
            continue

        if len(texts) == 1:
            line = texts[0]
            if line.startswith("Hist") and ": " in line and line.endswith(")"):
                current_category = line.split(": ", 1)[1].rsplit(" (", 1)[0]
            continue

        if not current_category or header_keys is None:
            continue
        if not re.fullmatch(r"\d{2}/\d{2}/\d{2}", texts[0]):
            continue

        padded = (texts + [""] * len(header_keys))[: len(header_keys)]
        row_map = {key: value for key, value in zip(header_keys, padded)}
        fatura, empresa = split_pipe(row_map.get("faturaemp"))
        _, doc_serie = split_pipe(row_map.get("docser"))

        counterpart = row_map.get("cliente") or row_map.get("fornecedor")
        rows.append(
            {
                "Tipo": "Receita" if kind == "receita" else "Despesa",
                "Categoria": current_category,
                "Vencimento": parse_date_br(row_map.get("vencto")),
                "Pagamento": parse_date_br(row_map.get("pagto")),
                "Fatura": fatura,
                "Empresa": empresa,
                "Fatura_Emp_Original": row_map.get("faturaemp"),
                "Baixa_na_Empresa": clean_text(row_map.get("baixanaempresa")),
                "Valor_Fatura": parse_currency(row_map.get("valorfatura")),
                "Juros": parse_currency(row_map.get("juros")),
                "Multa": parse_currency(row_map.get("multa")),
                "Desconto": parse_currency(row_map.get("desconto")),
                "Valor_Pago": parse_currency(row_map.get("valorpago")),
                "Contraparte": clean_text(counterpart),
                "Portador_Status": clean_text(row_map.get("portstatus")),
                "Doc_Cheque": clean_text(row_map.get("doccheque")),
                "Doc_Serie": clean_text(row_map.get("docser")),
                "Serie_Separada": doc_serie,
                "Origem": "Faturas recebidas" if kind == "receita" else "Faturas pagas",
                "Arquivo_Origem": path.name,
            }
        )

    return pd.DataFrame(rows)


def load_cashbook(path: Path) -> tuple[pd.DataFrame, int]:
    excel = pd.ExcelFile(path, engine="openpyxl")
    df = pd.read_excel(path, sheet_name=excel.sheet_names[0], engine="openpyxl")
    df.columns = [clean_text(col).replace("\n", " ") for col in df.columns]
    columns = {norm_key(col): col for col in df.columns}

    for key in ("debito", "credito", "saldo"):
        if key in columns:
            df[columns[key]] = pd.to_numeric(df[columns[key]], errors="coerce")

    history_col = columns["historico"]
    mask_invoice = df[history_col].fillna("").astype(str).str.contains(
        r"fatura\s*:", case=False, regex=True
    )
    filtered = df.loc[~mask_invoice].copy()
    filtered.insert(0, "Origem", "Livro caixa")
    return filtered, int(mask_invoice.sum())


def build_summary(
    receipts: pd.DataFrame,
    expenses: pd.DataFrame,
    unified: pd.DataFrame,
    cashbook_all_rows: int,
    cashbook_removed: int,
    cashbook_kept: int,
    receipt_files_found: int,
) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"Metrica": "Arquivos_receitas_localizados", "Valor": receipt_files_found},
            {
                "Metrica": "Arquivos_receitas_pendentes",
                "Valor": max(0, EXPECTED_RECEIPT_FILES - receipt_files_found),
            },
            {"Metrica": "Linhas_receitas_importadas", "Valor": len(receipts)},
            {"Metrica": "Categorias_receitas", "Valor": receipts["Categoria"].nunique()},
            {"Metrica": "Linhas_despesas_importadas", "Valor": len(expenses)},
            {"Metrica": "Categorias_despesas", "Valor": expenses["Categoria"].nunique()},
            {"Metrica": "Linhas_unificadas_total", "Valor": len(unified)},
            {"Metrica": "Linhas_totais_no_livro_caixa", "Valor": cashbook_all_rows},
            {"Metrica": "Linhas_removidas_do_livro_caixa_por_FATURA", "Valor": cashbook_removed},
            {"Metrica": "Linhas_mantidas_no_livro_caixa", "Valor": cashbook_kept},
        ]
    )


def autosize_and_format(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for column_cells in ws.columns:
            sample_values = [str(cell.value) if cell.value is not None else "" for cell in column_cells[:200]]
            max_length = max((len(value) for value in sample_values), default=0)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
                max(max_length + 2, 10), 40
            )

        headers = {norm_key(cell.value): cell.column for cell in ws[1] if cell.value}

        for key in ("vencimento", "pagamento", "dataemissao"):
            if key not in headers:
                continue
            for col in ws.iter_cols(
                min_col=headers[key], max_col=headers[key], min_row=2, max_row=ws.max_row
            ):
                for cell in col:
                    cell.number_format = "dd/mm/yyyy"

        for key in (
            "valorfatura",
            "juros",
            "multa",
            "desconto",
            "valorpago",
            "debito",
            "credito",
            "saldo",
        ):
            if key not in headers:
                continue
            for col in ws.iter_cols(
                min_col=headers[key], max_col=headers[key], min_row=2, max_row=ws.max_row
            ):
                for cell in col:
                    cell.number_format = "R$ #,##0.00"

    wb.save(workbook_path)


def main() -> None:
    paths = detect_paths()

    if not paths.cashbook.exists():
        raise FileNotFoundError(paths.cashbook)
    if not paths.expense_report.exists():
        raise FileNotFoundError(paths.expense_report)

    expenses = parse_html_report(paths.expense_report, kind="despesa")
    receipts = pd.concat(
        [parse_html_report(path, kind="receita") for path in paths.receipt_reports],
        ignore_index=True,
    ) if paths.receipt_reports else pd.DataFrame(columns=expenses.columns)

    unified = pd.concat([expenses, receipts], ignore_index=True)
    unified = unified.sort_values(
        by=["Pagamento", "Vencimento", "Tipo", "Categoria", "Fatura_Emp_Original"],
        na_position="last",
    ).reset_index(drop=True)

    cashbook, removed_from_cashbook = load_cashbook(paths.cashbook)
    cashbook = cashbook.sort_values(by=[cashbook.columns[1], cashbook.columns[2]]).reset_index(drop=True)

    summary = build_summary(
        receipts=receipts,
        expenses=expenses,
        unified=unified,
        cashbook_all_rows=len(cashbook) + removed_from_cashbook,
        cashbook_removed=removed_from_cashbook,
        cashbook_kept=len(cashbook),
        receipt_files_found=len(paths.receipt_reports),
    )

    with pd.ExcelWriter(paths.output_workspace, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="Resumo")
        unified.to_excel(writer, index=False, sheet_name="Movimentos_Unificados")
        cashbook.to_excel(writer, index=False, sheet_name="Livro_Caixa")

    autosize_and_format(paths.output_workspace)
    paths.output_desktop.write_bytes(paths.output_workspace.read_bytes())

    print(f"ARQUIVO_WORKSPACE: {paths.output_workspace}")
    print(f"ARQUIVO_DESKTOP: {paths.output_desktop}")
    print(f"ARQUIVOS_RECEITAS_LOCALIZADOS: {len(paths.receipt_reports)}")
    print(f"LINHAS_RECEITAS: {len(receipts)}")
    print(f"LINHAS_DESPESAS: {len(expenses)}")
    print(f"LINHAS_UNIFICADAS: {len(unified)}")
    print(f"LIVRO_CAIXA_MANTIDAS: {len(cashbook)}")
    print(f"LIVRO_CAIXA_REMOVIDAS_FATURA: {removed_from_cashbook}")


if __name__ == "__main__":
    main()
